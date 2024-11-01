import os
import pathlib
import re
import shutil
import time
import warnings
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import FreeSimpleGUI as sg
import fitz
import pandas as pd
from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

from helpers import (
    target_dict,
    and_regex,
    address_regex,
    date_regex,
    dollar_regex,
    postal_code_regex,
    ff,
    find_index,
    join_and_format_names,
    address_one_title_case,
    address_two_title_case,
    risk_address_title_case,
    unique_file_name,
    doc_types,
    content_pages,
    find_matching_paths,
    progressbar,
    format_postal_code,
    get_month_day,
    currency_to_float
)

warnings.simplefilter("ignore")

testing = []


def get_doc_types(doc):
    for doc_type in doc_types:
        page = doc[0]
        text_block = page.get_text("text", clip=doc_type.coordinates)
        if doc_type.keyword.casefold() in text_block.casefold():
            return doc_type.pdf_name


def get_content_pages(doc, pdf_name):
    pg_list = []
    for content_page in content_pages:
        keyword = content_page.keyword
        coordinates = content_page.coordinates
        for page_index in range(len(doc)):
            page = doc[page_index]
            # if the type of pdf found in get_doc_types matches the pdf_name
            if content_page.pdf_name == pdf_name:

                # Finds which page to stop on with coordinates being the starting range of pages
                if isinstance(keyword, str) and isinstance(coordinates, int):
                    if page.search_for(keyword):
                        for page_num in range(page_index + coordinates, page_index + 1):
                            pg_list.append(page_num + 1)

                # Finds which page contain a string match to a clipped area on the page
                elif isinstance(keyword, str) and isinstance(coordinates, tuple):
                    if page.search_for(keyword, clip=coordinates):
                        pg_list.append(page_index + 1)

                # Finds which page contain a regex match to a clipped area on the page
                elif isinstance(keyword, re.Pattern) and isinstance(coordinates, tuple):
                    for word_list in page.get_text("blocks", clip=coordinates):
                        if keyword.search(word_list[4]):
                            pg_list.append(page_index + 1)

    # If no search condition, return count of all pages
    if not pg_list:
        for page_num in range(doc.page_count):
            pg_list.append(page_num + 1)
    return pg_list


def search_for_input_dict(doc, pg_list):
    field_dict = {}
    for page_num in pg_list:
        page = doc[page_num - 1]
        wlist = page.get_text("blocks")
        text_boxes = [
            list(filter(None, inner_list[4].split("\n"))) for inner_list in wlist
        ]
        text_coords = [inner_list[:4] for inner_list in wlist]
        field_dict[page_num] = [
            [elem1, elem2] for elem1, elem2 in zip(text_boxes, text_coords)
        ]
    return field_dict


def append_word_to_dict(wlist, field_dict, append_duplicates):
    for words in wlist:
        word = words[4].strip().split("\n")
        if append_duplicates:
            field_dict.append(word)
        if word and word not in field_dict:
            field_dict.append(word)


def search_for_matches(doc, input_dict, type_of_pdf, target_dict):
    field_dict = defaultdict(lambda: defaultdict(list))
    try:
        targets = target_dict[type_of_pdf]
        for pg_num, pg in input_dict.items():
            page = doc[pg_num - 1]
            for i, wlist in enumerate(pg):
                for field_name, target in targets.items():
                    if target is not None:
                        try:
                            # This gets the name_and_address fields
                            if (
                                target.target_coordinates
                                and target.target_keyword is None
                            ):
                                tuple_list = doc[0].get_text(
                                    "blocks", clip=target.target_coordinates
                                )
                                append_word_to_dict(
                                    tuple_list,
                                    field_dict[type_of_pdf][field_name],
                                    target.append_duplicates,
                                )

                            # This gets the field using coordinates
                            elif (
                                target.target_keyword
                                and target.target_coordinates
                                and any(target.target_keyword in s for s in wlist[0])
                            ):
                                coords = tuple(
                                    x + y
                                    for x, y in zip(
                                        input_dict[pg_num][i][1],
                                        target.target_coordinates,
                                    )
                                )
                                word_list = page.get_text("blocks", clip=coords)
                                append_word_to_dict(
                                    word_list,
                                    field_dict[type_of_pdf][field_name],
                                    target.append_duplicates,
                                )

                            # This gets the field using keyword and indexing
                            elif isinstance(target.target_keyword, str) and any(
                                target.target_keyword in s for s in wlist[0]
                            ):
                                word = input_dict[pg_num][i + target.first_index][0][
                                    target.second_index
                                ]
                                if target.append_duplicates:
                                    field_dict[type_of_pdf][field_name].append(word)
                                elif target.join_list:
                                    field_dict[type_of_pdf][field_name].append(
                                        " ".join(word).split(", ")
                                    )
                                elif (
                                    word
                                    and word not in field_dict[type_of_pdf][field_name]
                                ):
                                    field_dict[type_of_pdf][field_name].append(word)

                            #  This gets the field using keyword (regex) and indexing
                            elif isinstance(target.target_keyword, re.Pattern):
                                word = input_dict[pg_num][i + target.first_index][0][
                                    target.second_index
                                ]
                                if (
                                    re.search(target.target_keyword, word)
                                    and word not in field_dict[type_of_pdf][field_name]
                                ):
                                    field_dict[type_of_pdf][field_name].append(word)
                        except IndexError:
                            continue
    except KeyError:
        print("This pdf dictionary has not been created yet")
    return field_dict


def format_named_insured(field_dict, dict_items, type_of_pdf):
    for name_and_address in dict_items["name_and_address"]:
        address_index = find_index(address_regex, name_and_address)

        if type_of_pdf == "Intact":
            # Join the list into a string so it can be processed
            name_string = " ".join(name_and_address[:address_index])

            # Split the string on '&' first to separate individuals
            individual_names = name_string.split("&")

            processed_names = []
            for name in individual_names:
                # Clean up spaces and split by commas
                name_parts = [part.strip().replace(":", "") for part in name.split(",")]

                if len(name_parts) == 2:
                    # Reverse "Last, First" to "First Last"
                    processed_names.append(f"{name_parts[1]} {name_parts[0]}".title())
                else:
                    # Already in "First Last" format, just strip and add
                    processed_names.append(name_parts[0].title())

            # Join names with ' & ' for the final output
            field_dict["named_insured"] = " and ".join(processed_names)

        else:
            # For non-"Intact" PDFs, handle differently
            names = re.sub(and_regex, "", ", ".join(name_and_address[:address_index]))
            field_dict["named_insured"] = join_and_format_names(
                names.split(", ")
            ).replace("  ", " ").replace(":", "")

    return field_dict


def format_insurer_name(field_dict, type_of_pdf):
    if (
        type_of_pdf == "Wawanesa"
        or type_of_pdf == "Intact"
        or type_of_pdf == "Family"
        or type_of_pdf == "Aviva"
    ):
        field_dict["insurer"] = type_of_pdf
    return field_dict


def format_mailing_address(field_dict, dict_items):
    for name_and_address in dict_items["name_and_address"]:
        pc_index = find_index(postal_code_regex, name_and_address)
        address_index = find_index(address_regex, name_and_address)
        city_province_p_code = " ".join(
            name_and_address[address_index + 1 : pc_index + 1]
        )
        if name_and_address[address_index : pc_index - 1] == []:
            field_dict["address_line_one"] = address_one_title_case(
                " ".join(name_and_address[address_index:pc_index])
            )
        else:
            field_dict["address_line_one"] = address_one_title_case(
                " ".join(name_and_address[address_index : pc_index - 1])
            )
        field_dict["address_line_two"] = address_two_title_case(
            re.sub(
                re.compile(r"Canada,"),
                "",
                re.sub(postal_code_regex, "", city_province_p_code),
            )
        )
        field_dict["address_line_three"] = (
            re.search(postal_code_regex, city_province_p_code).group().title()
        )
    return field_dict


def format_policy_number(field_dict, dict_items):
    if dict_items["policy_number"]:
        field_dict["policy_number"] = dict_items["policy_number"][0][0]
    return field_dict


def format_effective_date(field_dict, dict_items):
    if dict_items["effective_date"]:
        field_dict["effective_date"] = re.search(
            date_regex, dict_items["effective_date"][0][0]
        ).group()
    return field_dict


def sum_dollar_amounts(amounts):
    clean_amount_str = [
        a.replace("$", "")
        .replace(",", "")
        .replace(" 00", "")
        .replace(".00", "")
        .replace(" ", "")
        for a in amounts[0]
    ]

    def isDigit(x):
        try:
            float(x)
            return True
        except ValueError:
            return False
    total = sum(float(c) if isDigit(c) else 0 for c in clean_amount_str)
    return int(total)


def format_premium_amount(field_dict, dict_items):
    if dict_items["premium_amount"]:
        field_dict["premium_amount"] = "${:,.2f}".format(
            sum_dollar_amounts(dict_items["premium_amount"])
        )
    return field_dict


def format_additional_coverage(field_dict, dict_items, type_of_pdf):
    if (
        type_of_pdf == "Family"
        and dict_items["earthquake_coverage"]
        and re.search(dollar_regex, dict_items["earthquake_coverage"][0][0])
    ):
        field_dict["earthquake_coverage"] = True
    if type_of_pdf == "Aviva" or type_of_pdf == "Intact" or type_of_pdf == "Wawanesa":
        if dict_items["earthquake_coverage"]:
            field_dict["earthquake_coverage"] = True
    if type_of_pdf == "Intact" and dict_items["ground_water"]:
        field_dict["ground_water"] = True
    if type_of_pdf == "Wawanesa" and dict_items["tenant_vandalism"]:
        field_dict["tenant_vandalism"] = True
    if dict_items["overland_water"]:
        field_dict["overland_water"] = True
    if dict_items["service_line"]:
        field_dict["service_line"] = True
    if type_of_pdf == "Family" and dict_items["overland_water"]:
        field_dict["service_line"] = True


def find_risk_addresses(risk_addresses):
    matched = []
    for risk_address1 in risk_addresses:
        for risk_address in risk_address1:
            if (
                re.search(postal_code_regex, risk_address)
                and risk_address not in matched
            ):
                matched.append(risk_address)
    return matched


def format_risk_address(field_dict, dict_items):
    risk_addresses = find_risk_addresses(dict_items["risk_address"])
    for index, risk_address in enumerate(risk_addresses):
        field_dict[f"risk_address_{index + 1}"] = risk_address_title_case(
            re.sub(postal_code_regex, "", risk_addresses[index]).rstrip(", ")
        )
    return field_dict


def format_form_type(field_dict, dict_items, type_of_pdf):
    for form_types in dict_items["form_type"]:
        for index, form_type in enumerate(form_types):
            if "comprehensive".casefold() in form_type.casefold():
                field_dict[f"form_type_{index + 1}"] = "Comprehensive Form"
            if "broad".casefold() in form_type.casefold():
                field_dict[f"form_type_{index + 1}"] = "Broad Form"
            if "basic".casefold() in form_type.casefold():
                field_dict[f"form_type_{index + 1}"] = "Basic Form"
            if "fire & extended".casefold() in form_type.casefold():
                field_dict[f"form_type_{index + 1}"] = "Fire + EC"
            if "DOLCE VITA".casefold() in form_type.casefold():
                field_dict[f"form_type_{index + 1}"] = "Comprehensive Form"
            if type_of_pdf == "Family":
                if "included".casefold() in form_type.casefold():
                    field_dict[f"form_type_{index + 1}"] = "Comprehensive Form"
                else:
                    field_dict[f"form_type_{index + 1}"] = "Specified Perils"
    return field_dict


def format_risk_type(field_dict, dict_items, type_of_pdf):
    for risk_types in dict_items["risk_type"]:
        for index, risk_type in enumerate(risk_types):
            if "seasonal".casefold() in risk_type.casefold():
                field_dict["seasonal"] = True
            if "home".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "home"
            elif (
                type_of_pdf == "Aviva"
                and "condominium".casefold() in risk_type.casefold()
            ):
                field_dict[f"risk_type_{index + 1}"] = "condo"
            elif type_of_pdf == "Family" and "Condominium".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "condo"
            elif (
                type_of_pdf == "Intact"
                and "condominium ".casefold() in risk_type.casefold()
            ):
                field_dict[f"risk_type_{index + 1}"] = "condo"
                field_dict["condo_deductible_1"] = "$100,000"
            elif (
                type_of_pdf == "Intact"
                and "Rented Condominium".casefold() in risk_type.casefold()
            ):
                field_dict[f"risk_type_{index + 1}"] = "rented_condo"
                field_dict["condo_deductible_1"] = "$100,000"
            elif type_of_pdf == "Wawanesa" and "Rental Condominium" in risk_type:
                field_dict[f"risk_type_{index + 1}"] = "rented_condo"
            elif type_of_pdf == "Wawanesa" and "Condominium" in risk_type:
                field_dict[f"risk_type_{index + 1}"] = "condo"
            elif "rented dwelling".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "rented_dwelling"
            elif "revenue".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "rented_dwelling"
            elif "rental".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "rented_condo"
            elif "tenant".casefold() in risk_type.casefold():
                field_dict[f"risk_type_{index + 1}"] = "tenant"

    return field_dict


def format_number_families(field_dict, dict_items, type_of_pdf):
    keywords = {
        "One": 1,
        "Two": 2,
        "Three": 3,
        "1": 1,
        "2": 2,
        "3": 3,
        "Additional Family": 2,
        "002 Additional Family": 3,
    }
    if (
        type_of_pdf == "Wawanesa"
        or type_of_pdf == "Intact"
        or type_of_pdf == "Family"
        or type_of_pdf == "Aviva"
    ):
        if not dict_items["number_of_families"]:
            field_dict["number_of_families_1"] = keywords.get("1", None)
    for families in dict_items["number_of_families"]:
        for index, number_of_families in enumerate(families):
            if type_of_pdf == "Aviva":
                field_dict[f"number_of_families_{index + 1}"] = keywords.get(
                    number_of_families, None
                )
            match = re.search(r"\b(\d+)\b", number_of_families)
            if type_of_pdf == "Family" and match:
                field_dict[f"number_of_families_{index + 1}"] = keywords.get(
                    str(int(match.group(1)) + 1), None
                )
            if type_of_pdf == "Intact" or type_of_pdf == "Wawanesa":
                field_dict[f"number_of_families_{index + 1}"] = keywords.get(
                    number_of_families, None
                )
    if not dict_items["number_of_families"] and dict_items["number_of_units"]:
        for families in dict_items["number_of_units"]:
            for index, number_of_units in enumerate(families):
                field_dict[f"number_of_families_{index + 1}"] = keywords.get(
                    number_of_units, None
                )
    return field_dict


def format_condo_deductible(field_dict, dict_items, type_of_pdf):
    for deductibles in dict_items["condo_deductible"]:
        if type_of_pdf == "Family":
            field_dict["condo_deductible_1"] = re.search(
                dollar_regex, deductibles[0]
            ).group()
            if len(deductibles) > 1:
                if re.search(dollar_regex, deductibles[1]):
                    field_dict["condo_earthquake_deductible_1"] = re.search(
                        dollar_regex, deductibles[1]
                    ).group()
            else:
                if re.search(dollar_regex, deductibles[0]):
                    field_dict["condo_earthquake_deductible_1"] = re.search(
                        dollar_regex, deductibles[0]
                    ).group()
        for index, condo_deductible in enumerate(deductibles):
            if dict_items["condo_deductible"] and type_of_pdf != "Family":
                field_dict[f"condo_deductible_{index + 1}"] = re.search(
                    dollar_regex, condo_deductible
                ).group()
            if type_of_pdf == "Aviva":
                field_dict["condo_earthquake_deductible_1"] = re.search(
                    dollar_regex, condo_deductible
                ).group()
    return field_dict


def format_condo_earthquake_deductible(field_dict, dict_items, type_of_pdf):
    if (
        type_of_pdf == "Intact"
        and dict_items["earthquake_coverage"]
        and not dict_items["condo_earthquake_deductible"]
    ):
        field_dict["condo_earthquake_deductible_1"] = "$2,500"
    for deductibles in dict_items["condo_earthquake_deductible"]:
        for index, condo_earthquake_deductible in enumerate(deductibles):
            if type_of_pdf == "Intact" and dict_items["condo_earthquake_deductible"]:
                field_dict["condo_earthquake_deductible_1"] = "$25,000"
            else:
                field_dict[f"condo_earthquake_deductible_{index + 1}"] = re.search(
                    dollar_regex, condo_earthquake_deductible
                ).group()
    return field_dict


def format_policy(flattened_dict, type_of_pdf):
    field_dict = {}
    if (
        type_of_pdf
        and type_of_pdf == "Wawanesa"
        or type_of_pdf == "Intact"
        or type_of_pdf == "Family"
        or type_of_pdf == "Aviva"
    ):
        format_named_insured(field_dict, flattened_dict, type_of_pdf)
        format_insurer_name(field_dict, type_of_pdf)
        format_mailing_address(field_dict, flattened_dict)
        format_policy_number(field_dict, flattened_dict)
        format_effective_date(field_dict, flattened_dict)
        format_premium_amount(field_dict, flattened_dict)
        format_additional_coverage(field_dict, flattened_dict, type_of_pdf)
        format_risk_address(field_dict, flattened_dict)
        format_form_type(field_dict, flattened_dict, type_of_pdf)
        format_risk_type(field_dict, flattened_dict, type_of_pdf)
        format_number_families(field_dict, flattened_dict, type_of_pdf)
        format_condo_deductible(field_dict, flattened_dict, type_of_pdf)
        format_condo_earthquake_deductible(field_dict, flattened_dict, type_of_pdf)
    return field_dict


def create_pandas_df(data_dict):
    df = pd.DataFrame([data_dict])
    df["today"] = datetime.today().strftime("%B %d, %Y")
    try:
        df["effective_date"] = pd.to_datetime(df["effective_date"]).dt.strftime(
            "%B %d, %Y"
        )
        expiry_date = pd.to_datetime(df["effective_date"]) + pd.offsets.DateOffset(
            years=1
        )
        df["expiry_date"] = expiry_date.dt.strftime("%B %d, %Y")
    except KeyError:
        print("invalid effective_date")
    return df


def write_to_new_docx(docx, rows):
    template_path = base_dir / "assets" / docx
    doc = DocxTemplate(template_path)
    doc.render(rows)
    output_path = output_dir / f"{rows["named_insured"].rstrip(".").rstrip(":")}.docx"
    doc.save(unique_file_name(output_path))


# def write_to_pdf(pdf, dictionary, rows):
#     pdf_path = (base_dir / "templates" / pdf)
#     output_path = base_dir / "output" / f"{rows["named_insured"]} {rows["risk_type"].title()}.pdf"
#     output_path.parent.mkdir(exist_ok=True)
#     reader = PdfReader(pdf_path)
#     writer = PdfWriter()
#     for page_num in range(len(reader.pages)):
#         page = reader.pages[page_num]
#         writer.add_page(page)
#         writer.updatePageFormFieldValues(page, dictionary)
#     with open(unique_file_name(output_path), "wb") as output_stream:
#         writer.write(output_stream)


def sort_renewal_list():
    xlsx_files = Path(input_dir).glob("*.xlsx")
    xls_files = Path(input_dir).glob("*.xls")
    files = list(xlsx_files) + list(xls_files)
    all_dfs = []
    if len(files) > 0:
        for file in progressbar(files, prefix="Progress: ", size=40):
            df = (
                pd.read_excel(file, engine="xlrd")
                if file.suffix == ".XLS"
                else pd.read_excel(file, engine="openpyxl")
            )
            all_dfs.append(df)

        df = pd.concat(all_dfs, ignore_index=True)

        output_path = output_dir / "renewal_list.xlsx"
        output_path = unique_file_name(output_path)
        try:
            column_list = [
                "policynum",
                "ccode",
                "name",
                "pcode",
                "csrcode",
                "insurer",
                "buscode",
                "renewal",
                "Pulled",
                "D/L",
                # "agen_dir",
                # "prem_amt",
                # "h_postzip"
            ]

            df = df.reindex(columns=column_list)
            df = df.drop_duplicates(subset=["policynum"], keep=False)
            df["renewal_1"] = pd.to_datetime(df["renewal"], dayfirst=True).dt.strftime(
                "%d-%b"
            )
            df["renewal"] = pd.to_datetime(df["renewal"], dayfirst=True).dt.strftime(
                "%m%d"
            )
            df.sort_values(
                ["insurer", "renewal", "name"],
                ascending=[True, True, True],
                inplace=True,
            )
            df["renewal"] = df["renewal_1"]
            df = df.drop("renewal_1", axis=1)
            list_with_spaces = []
            for x, y in df.groupby("insurer", sort=False):
                list_with_spaces.append(y)
                list_with_spaces.append(
                    pd.DataFrame([[float("NaN")] * len(y.columns)], columns=y.columns)
                )
            df = pd.concat(list_with_spaces, ignore_index=True).iloc[:-1]
            # df = df[((df['pcode'] == 'FS') | (df['csrcode'] == 'FS')) & (df['buscode'] == 'COMM')]
            # print(df)
            if not os.path.isfile(output_path):
                writer = pd.ExcelWriter(output_path, engine="openpyxl")
            else:
                writer = pd.ExcelWriter(
                    output_path, mode="a", if_sheet_exists="replace", engine="openpyxl"
                )

            df.to_excel(writer, sheet_name="Sheet1", index=False)
            writer.close()
            # Load the workbook and get the sheet
            wb = load_workbook(output_path)
            ws = wb.active

            # Set font size and alignment for all cells
            for row in ws.iter_rows():
                for cell in row:
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(horizontal="left")

            # Define the reference for the table
            ref = f"A1:{chr(65 + df.shape[1] - 1)}{df.shape[0] + 1}"

            # Create an Excel table
            tab = Table(displayName="Table1", ref=ref)
            style = TableStyleInfo(
                name="TableStyleLight1",  # This style alternates grey for each row
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            tab.tableStyleInfo = style

            # Add the table to the worksheet
            ws.add_table(tab)

            # Adjust column widths for specified columns
            for i, col in enumerate(column_list, 1):
                max_length = max(df[col].astype(str).map(len).max(), len(col)) + 4

                med_length = max(df[col].astype(str).map(len).max(), len(col)) + 2.5
                min_length = max(df[col].astype(str).map(len).max(), len(col))
                if col in ["pcode", "csrcode", "Pulled", "D/L"]:
                    ws.column_dimensions[chr(64 + i)].width = (
                        5.0  # Set fixed width for Pulled and D/L columns
                    )
                elif col in ["ccode"]:
                    ws.column_dimensions[chr(64 + i)].width = max_length
                elif col in ["policynum"]:
                    ws.column_dimensions[chr(64 + i)].width = med_length
                else:
                    ws.column_dimensions[chr(64 + i)].width = min_length

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col in ["Pulled", "D/L"]:
                col_index = column_list.index(col) + 1
                for row in range(1, df.shape[0] + 2):  # +2 to include the header
                    ws.cell(row=row, column=col_index).border = border

            # Set the first row to repeat on each printed page
            ws.print_title_rows = "1:1"

            # Set print area to fit all columns on one page
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToPage = True

            # Set margins
            ws.page_margins = PageMargins(
                top=1.91 / 2.54,  # Convert cm to inches
                bottom=1.91 / 2.54,  # Convert cm to inches
                left=1.78 / 2.54,  # Convert cm to inches
                right=0.64 / 2.54,  # Convert cm to inches
            )

            # Save the workbook
            wb.save(output_path)
        except TypeError:
            return
    if len(files) == 0:
        print("Missing renewal list in 'Input' folder.")
    else:
        print(f"******** Sort Renewal List ran successfully ********")


def renewal_letter(excel_path1):
    excel_folder = base_dir / "assets"
    xlsx_files = excel_folder.rglob("*.xlsx")
    xls_files = excel_folder.rglob("*.xls")
    all_dfs = []
    files = list(xlsx_files) + list(xls_files)
    def get_glass_policy():
        for file in files:
            df = (
                pd.read_excel(file, engine="xlrd")
                if file.suffix == ".XLS"
                else pd.read_excel(file, engine="openpyxl")
            )
            all_dfs.append(df)
            df = pd.concat(all_dfs, ignore_index=True)
            df["postal_code"] = df["h_postzip"].apply(format_postal_code)
            return df
    df_glass = get_glass_policy()
    if df_glass is not None:
        df_glass = df_glass.drop_duplicates(subset=["policynum"], keep=False)
    doc_found = 0
    pdf_files = input_dir.glob("*.pdf")
    for pdf in progressbar(list(pdf_files), prefix="Progress: ", size=40):
        with fitz.open(pdf) as doc:
            # print(
            #     f"\n<==========================>\n\nFilename is: {Path(pdf).stem}{Path(pdf).suffix} "
            # )
            doc_type = get_doc_types(doc)
            # print(f"This is a {doc_type} policy.")
            if doc_type:
                doc_found += 1
                pg_list = get_content_pages(doc, doc_type)
                input_dict = search_for_input_dict(doc, pg_list)
                dict_items = search_for_matches(doc, input_dict, doc_type, target_dict)
                try:
                    formatted_dict = format_policy(ff(dict_items[doc_type]), doc_type)
                except AttributeError:
                    if doc_type == "Family":
                        try:
                            formatted_dict = format_policy(
                                ff(
                                    search_for_matches(
                                        doc,
                                        search_for_input_dict(
                                            doc, get_content_pages(doc, "Family_Legal")
                                        ),
                                        "Family_Legal",
                                        target_dict,
                                    )["Family_Legal"]
                                ),
                                "Family",
                            )
                        except AttributeError:
                            print("Problem reading pdf")
                try:
                    df = create_pandas_df(formatted_dict)
                except KeyError:
                    continue
                df["broker_name"] = pd.read_excel(
                    excel_path1, sheet_name=0, header=None
                ).at[6, 1]
                df["on_behalf"] = pd.read_excel(
                    excel_path1, sheet_name=0, header=None
                ).at[8, 1]
                if doc_type and doc_type != "ICBC":
                    if df_glass is not None:

                        # Create a new column for glass_policynum
                        df["glass_policynum"] = None
                        df["glass"] = False
                        # Perform row-by-row comparison
                        for _, row in df_glass.iterrows():
                            if row["insurer"] == "REL":  # Check if insurer is "REL"
                                for i, row2 in df.iterrows():
                                    if row2["risk_type_1"] == "home":

                                        if row["postal_code"] == row2["address_line_three"]:
                                            if get_month_day(row["renewal"]) == get_month_day(row2["expiry_date"]):
                                                df.at[i, "glass_policynum"] = row["policynum"]
                                                if row["policynum"]:
                                                    df.at[i, "glass"] = True
                                                df.at[
                                                    i, "premium_amount"] = f"${currency_to_float(row2['premium_amount']) + float(row['prem_amt']):.2f}"
                    for rows in df.to_dict(orient="records"):
                        write_to_new_docx("Renewal Letter New.docx", rows)

    if doc_found > 0:
        print(f"******** Auto Renewal Letter ran successfully ********")
    else:
        print("Missing insurance policy documents in 'Input' folder")


def renewal_letter_manual(excel_data1):

    df = pd.DataFrame([excel_data1])
    df["today"] = datetime.today().strftime("%B %d, %Y")
    df["mailing_address"] = (
        df[["address_line_one", "address_line_two"]]
        .astype(str)
        .apply(lambda x: ", ".join(x[:-1]) + " " + x[-1:], axis=1)
    )
    df["risk_address_1"] = df["risk_address_1"].fillna(df["mailing_address"])
    df["effective_date"] = pd.to_datetime(df["effective_date"]).dt.strftime("%B %d, %Y")
    # print(df)
    for rows in progressbar(df.to_dict(orient="records"), prefix="Progress: ", size=40):
        write_to_new_docx("Renewal Letter New.docx", rows)
    print(f"******** Manual Renewal Letter ran successfully ********")


def get_excel_data(excel_path1):
    data = {}
    try:
        df = pd.read_excel(excel_path1, sheet_name=0, header=None)
        data["broker_name"] = df.at[6, 1]
        data["on_behalf"] = df.at[8, 1]
        data["risk_type_1"] = df.at[12, 1]
        data["named_insured"] = df.at[14, 1]
        data["insurer"] = df.at[15, 1]
        data["policy_number"] = df.at[16, 1]
        data["effective_date"] = df.at[17, 1]
        data["address_line_one"] = df.at[19, 1]
        data["address_line_two"] = df.at[20, 1]
        data["address_line_three"] = df.at[21, 1]
        data["risk_address_1"] = df.at[23, 1]
    except KeyError:
        return None
    return data


def format_icbc(dict_items, type_of_pdf):
    field_dict = {}
    if type_of_pdf and type_of_pdf == "ICBC":
        if not dict_items["licence_plate"]:
            field_dict["licence_plate"] = "NONLIC"
        if not dict_items["transaction_type"]:
            field_dict["transaction_type"] = "PLACEHOLDER"
        for license_plates in dict_items["licence_plate"]:
            plate_number = re.sub(
                re.compile(r"Licence Plate Number "), "", license_plates[0]
            )
            field_dict["licence_plate"] = plate_number
        for transaction_types in dict_items["transaction_type"]:
            transaction = re.sub(
                re.compile(r"Transaction Type "), "", transaction_types[0]
            )
            field_dict["transaction_type"] = transaction
        for cancellations in dict_items["cancellation"]:
            for index, cancellation in enumerate(cancellations):
                if isinstance(cancellation, str):
                    field_dict["transaction_type"] = "CANCEL"
        for storages in dict_items["storage"]:
            for index, storage in enumerate(storages):
                if isinstance(storage, str):
                    field_dict["transaction_type"] = "STORAGE"
        for tops in dict_items["top"]:
            for index, top in enumerate(tops):
                if isinstance(top, str):
                    field_dict["transaction_type"] = "TOP"
        for name_codes in dict_items["name_code"]:
            for index, name_code in enumerate(name_codes):
                name = re.search(re.compile(r"(?<= - )\b\w{2,3}\b(?= - )"), name_code)
                if name is not None:
                    field_dict["name_code"] = name.group().upper()
                else:
                    field_dict["name_code"] = "HOUSE"
        for transaction_timestamps in dict_items["transaction_timestamp"]:
            transaction1 = re.sub(
                re.compile(r"Transaction Timestamp "), "", transaction_timestamps[0]
            )
            field_dict["transaction_timestamp"] = transaction1
        for insured_names in dict_items["owner_name"]:
            field_dict["insured_name"] = insured_names[0].rstrip(".")
        for insured_names in dict_items["applicant_name"]:
            field_dict["insured_name"] = insured_names[0].rstrip(".")
        for insured_names in dict_items["insured_name"]:
            field_dict["insured_name"] = insured_names[0].rstrip(".")
        format_named_insured(field_dict, dict_items, type_of_pdf)
    return field_dict


def icbc_filename(df):
    if df["licence_plate"].at[0] == "NONLIC" and df["transaction_type"].at[0] not in [
        "RENEW",
        "NEW",
        "CHANGE",
        "CANCEL",
        "STORAGE",
        "TOP",
    ]:
        return f"{df['insured_name'].at[0].title()}.pdf"
    elif df["licence_plate"].at[0] == "NONLIC" and df["transaction_type"].at[0] not in [
        "CANCEL"
    ]:
        return f"{df['insured_name'].at[0].title()} {df['transaction_type'].at[0].title()}.pdf"
    elif df["transaction_type"].at[0] not in [
        "RENEW",
        "NEW",
        "STORAGE",
        "TOP",
    ]:
        return f"{df['licence_plate'].at[0]} {df['transaction_type'].at[0].title()}.pdf"
    else:
        return f"{df['licence_plate'].at[0]}.pdf"


def get_icbc_doc_types(doc):
    for doc_type in doc_types:
        page = doc[0]
        text_block = page.get_text("text", clip=doc_type.coordinates)
        if doc_type.keyword.casefold() in text_block.casefold():
            return doc_type.pdf_name


def search_for_icbc_input_dict(doc):
    field_dict = {}
    for page_num in range(len(doc)):
        page = doc[page_num - 1]
        wlist = page.get_text("blocks")
        text_boxes = [
            list(filter(None, inner_list[4].split("\n"))) for inner_list in wlist
        ]
        text_coords = [inner_list[:4] for inner_list in wlist]
        field_dict[page_num] = [
            [elem1, elem2] for elem1, elem2 in zip(text_boxes, text_coords)
        ]
    return field_dict


def rename_icbc(drive_letter, number_of_pdfs, names, icbc_folder_name):
    global testing
    loop_counter = 0
    scan_counter = 0
    copy_counter = 0
    # icbc_output_directory.mkdir(exist_ok=True)
    icbc_input_directory = Path.home() / "Downloads"
    icbc_output_directory = (
        f"{drive_letter}:\\{icbc_folder_name}"
        if not testing
        else Path.home() / "Desktop" / "ICBC Copies"
    )
    if testing:
        icbc_output_directory.mkdir(exist_ok=True)
    if not testing and not Path(f"{drive_letter}:\\").exists():
        print(
            "Change the drive letter in 'input.xlsx' to the same one as the 'Shared' drive."
        )
        # os.system("pause")
        # return
    if not testing and not Path(icbc_output_directory).exists():
        print("Check if the ICBC folder name is correct.")
        # os.system("pause")
        # return

    pdf_files1 = list(icbc_input_directory.glob("*.pdf"))
    pdf_files1 = sorted(
        pdf_files1, key=lambda file: pathlib.Path(file).lstat().st_mtime, reverse=True
    )
    paths = list(Path(icbc_output_directory).rglob("*.pdf"))
    file_names = [path.stem.split()[0] for path in paths]
    processed_timestamps = set()
    for pdf in progressbar(pdf_files1[:number_of_pdfs], prefix="Progress: ", size=40):
        loop_counter += 1
        with fitz.open(pdf) as doc:
            pp = False
            pp_block = doc[0].get_text(
                "text",
                clip=(
                    425.40240478515625,
                    35.96635437011719,
                    557.9161376953125,
                    48.300140380859375,
                ),
            )
            if "Payment Plan Agreement".casefold() in pp_block.casefold():
                pp = True
            doc_type = get_icbc_doc_types(doc)
            if doc_type and not pp and doc_type == "ICBC":
                scan_counter += 1
                input_dict = search_for_icbc_input_dict(doc)
                dict_items = search_for_matches(doc, input_dict, doc_type, target_dict)
                formatted_dict = format_icbc(ff(dict_items[doc_type]), doc_type)

                try:
                    df = pd.DataFrame([formatted_dict])
                except KeyError:
                    continue
                timestamp = int(df["transaction_timestamp"].at[0])
                if timestamp in processed_timestamps:
                    continue
                processed_timestamps.add(timestamp)
                icbc_file_name = icbc_filename(df)
                icbc_output_dir = (
                    Path(icbc_output_directory)
                    if df["name_code"].at[0].upper() == "HOUSE"
                    or names.get(df["name_code"].at[0].upper()) is None
                    else Path(
                        f"{icbc_output_directory}/{names.get(df['name_code'].at[0].upper())}"
                    )
                )

                if testing:
                    icbc_output_dir.mkdir(exist_ok=True)

                icbc_output_path = icbc_output_dir / icbc_file_name
                if not testing and not Path(icbc_output_dir).exists():
                    print(
                        "The producer folder does not exists, check if mappings in 'input.xlsx' is correct."
                    )
                    # os.system("pause")
                    # return

                target_filename = Path(icbc_file_name).stem.split()[0]
                matching_transaction_ids = []
                if target_filename in file_names:
                    # Find matching paths for the target filename
                    matching_paths = find_matching_paths(target_filename, paths)
                    for path_name in matching_paths:
                        with fitz.open(path_name) as doc1:
                            target_transaction_id = doc1[0].get_text(
                                "text",
                                clip=(
                                    409.97900390625,
                                    63.84881591796875,
                                    576.0,
                                    72.82147216796875,
                                ),
                            )
                            if target_transaction_id:
                                if target_transaction_id:
                                    match = re.match(r".*?(\d+)", target_transaction_id)
                                    if match:  # Check if the match was successful
                                        temp_match = int(match.group(1))
                                        matching_transaction_ids.append(temp_match)

                if timestamp not in matching_transaction_ids:
                    try:
                        shutil.copy(pdf, unique_file_name(icbc_output_path))
                        copy_counter += 1
                    except FileNotFoundError:
                        print(
                            "*****************Correct the above errors.*****************"
                        )
    if scan_counter > 0:
        print(f"Scanned: {scan_counter} out of {loop_counter} documents")
        print(f"Copied: {copy_counter} out of {scan_counter} documents")
    else:
        print("There are no policy documents in the Downloads folder")


def find_pages_with_text(doc, search_text, not_search_text):
    pages = []
    for page_index in range(len(doc)):
        page = doc[page_index]
        text = page.get_text("text")
        if not search_text and not not_search_text and not text:
            pages.append(page_index)
        elif search_text in text and not_search_text not in text:
            pages.append(page_index)
    return pages


def find_blank_pages_no_images(doc):
    pages = []
    for page_index in range(len(doc)):
        page = doc[page_index]
        text = page.get_text("text")
        pix = page.get_images()
        if not pix and not text:
            pages.append(page_index)
    return pages


def delete_intact_broker_copies():
    doc_found = 0
    pdf_files = input_dir.glob("*.pdf")
    for pdf in progressbar(list(pdf_files), prefix="Progress: ", size=40):
        with fitz.open(pdf) as doc:
            # print(
            #     f"\n<==========================>\n\nFilename is: {Path(pdf).stem}{Path(pdf).suffix} "
            # )
            intact_doc = get_doc_types(doc)

            # print(f"This is a {intact_doc} policy.")
            if intact_doc == "Intact":
                doc_found += 1
                broker_pages = find_pages_with_text(
                    doc, "BROKER COPY", "Property Summary"
                )
                # logo_pages = find_pages_with_text(doc, "KMJ", "BROKER COPY")
                blank_logo_pages = find_pages_with_text(doc, "", "BROKER COPY")
                blank_pages = find_blank_pages_no_images(doc)
                mortgage_pages = find_pages_with_text(doc, "MORTGAGE", "BROKER COPY")
                stat_cond_pages = find_pages_with_text(
                    doc, "STATUTORY CONDITIONS", "KMJ"
                )
                pages_to_remove = set(broker_pages + blank_pages + mortgage_pages)
                # for page_num in broker_pages:
                #     if page_num + 1 in logo_pages:
                #         pages_to_remove.add(page_num + 1)
                for page_num in broker_pages:
                    if page_num + 1 in blank_logo_pages:
                        pages_to_remove.add(page_num + 1)
                for page_num in mortgage_pages:
                    if page_num + 1 in stat_cond_pages:
                        pages_to_remove.add(page_num + 1)
                pages_to_remove = sorted(pages_to_remove, reverse=True)
                output_path = output_dir / f"{Path(pdf).stem}{Path(pdf).suffix}"
                # print("*************  Preparing Intact Customer Copy *************")
                doc.delete_pages(pages_to_remove)
                doc.save(unique_file_name(output_path), garbage=4, deflate=True)
    if doc_found > 0:
        print(f"******** Delete Intact Broker/Mortgage Pages ran successfully ********")
    else:
        print("Missing Intact policy documents in the 'Input' folder")


base_dir = Path(__file__).parent.parent
input_dir = Path.home() / "Desktop" / "Input (this folder can be deleted)"
output_dir = Path.home() / "Desktop"

font = ("Arial", 11)
font2 = ("Arial", 8)
button_size = (18, 1)
layout = [
    [
        sg.Multiline(
            size=(47, 4),
            echo_stdout_stderr=True,
            reroute_stdout=True,
            autoscroll=True,
            background_color="LightYellow",
            text_color="indigo",
            key="-MLINE-",
            font=font2,
        ),
        sg.Column(
            [
                [
                    sg.Button(
                        "Sort Renewal List",
                        font=font,
                        button_color="MediumSeaGreen",
                        size=button_size,
                    ),
                ],
                [
                    sg.Button(
                        "Intact Customer Copy",
                        font=font,
                        button_color="MediumSeaGreen",
                        size=button_size,
                    )
                ],
            ],
            background_color="LightYellow",
            element_justification="right",
        ),
    ],
    [
        sg.Exit(s=16, button_color="tomato", font=font),
        sg.Button("Auto Renewal Letter", font=font, button_color="MediumSeaGreen"),
        sg.Button(
            "Copy Rename ICBC", font=font, button_color="Indigo", size=button_size
        ),
    ],
]
window = sg.Window(
    "GUI Interface using Visual Basic", layout, background_color="LightYellow"
)


def main():
    excel_path = base_dir / "input.xlsx"  # name of Excel
    excel_data = get_excel_data(excel_path)
    df_excel = pd.read_excel(excel_path, sheet_name=0, header=None)
    names = {}
    icbc_folder_name = df_excel.at[31, 1]
    for i in range(33, 49):
        names[df_excel.at[i, 1]] = df_excel.at[i, 2]
    task = df_excel.at[2, 1]
    drive_letter = df_excel.at[29, 1]
    number_of_pdfs = int(df_excel.at[27, 1])
    if not Path(input_dir).exists():
        input_dir.mkdir(exist_ok=True)
        return
    if task == "GUI Interface using Visual Basic":
        while True:
            event, values = window.read()
            if event in (sg.WINDOW_CLOSED, "Exit"):
                break
            if event == "Auto Renewal Letter":
                renewal_letter(excel_path)
                time.sleep(3)
            if event == "Sort Renewal List":
                sort_renewal_list()
                time.sleep(3)
            if event == "Intact Customer Copy":
                delete_intact_broker_copies()
                time.sleep(3)
            if event == "Copy Rename ICBC":
                rename_icbc(drive_letter, number_of_pdfs, names, icbc_folder_name)
                time.sleep(3)
        window.close()
    elif task == "Auto Renewal Letter":
        # output_dir.mkdir(exist_ok=True)
        renewal_letter(excel_path)
    elif task == "Manual Renewal Letter":
        # output_dir.mkdir(exist_ok=True)
        renewal_letter_manual(excel_data)
    elif task == "Sort Renewal List":
        # output_dir.mkdir(exist_ok=True)
        sort_renewal_list()
    elif task == "Copy/Rename ICBC Transactions":
        rename_icbc(drive_letter, number_of_pdfs, names, icbc_folder_name)
    elif task == "Delete Intact Broker/Mortgage Pages":
        # output_dir.mkdir(exist_ok=True)
        delete_intact_broker_copies()


if __name__ == "__main__":
    main()
